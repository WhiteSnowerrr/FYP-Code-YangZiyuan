#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Oct 29 01:51:48 2023

@author: yangziyuan
"""

import pandas as pd
import numpy as np
import math as m

#%% rewrite the output function to chose the output path
def newgetNewFileName(name, ext, path=''):
    from pathlib import Path
    fileName = name + '.' + ext
    theFile = Path(path+fileName)
    number = int(0)
    while theFile.is_file():
        fileName = f'{name}~{number:02d}.{ext}'
        theFile = Path(path+fileName)
        number += 1
    return fileName
def newwriteHtml(t, onlyRobust=True, path=''):
    """Write the results in an HTML file."""
    t.data.htmlFileName = newgetNewFileName(t.data.modelName, 'html', path)
    fileplace = path+t.data.htmlFileName
    with open(fileplace, 'w', encoding='utf-8') as f:
        f.write(t.getHtml(onlyRobust))

def writeExcel(t, onlyRobust=True, path=''):
    import datetime
    import biogeme.version as bv
    fileplace = path+newgetNewFileName(t.data.modelName, 'xlsx', path)
    writer = pd.ExcelWriter(fileplace)
    # Estimation report
    index = []
    temp = []
    index.append(' ')
    temp.append(' ')
    
    a,b = getXlsxHeader()
    index += a
    temp += b
    
    index.append(' ')
    temp.append(' ')
    index.append('Xlsx file generated time')
    temp.append(datetime.datetime.now())
    index.append('Report file')
    temp.append(newgetNewFileName(t.data.modelName, 'xlsx', path))
    index.append('Database name')
    temp.append(t.data.dataname)
    index.append(' ')
    temp.append(' ')
    if np.abs(t.data.smallestEigenValue) <= t.identification_threshold:
            index.append('Warning: identification issue')
            t0 = (f'The second derivatives matrix is close to singularity. '
            f'The smallest eigenvalue is '
            f'{np.abs(t.data.smallestEigenValue):.3g}. This warning is '
            f'triggered when it is smaller than the parameter '
            f'identification_threshold = '
            f'{t.identification_threshold}.\n'
            f'Variables involved:\n')
            for i, ev in enumerate(t.data.smallestEigenVector):
                if np.abs(ev) > t.identification_threshold:
                    t0 += (
                        f'{ev:.3g}'
                        f' * '
                        f'{t.data.betaNames[i]}\n'
                    )
            temp.append(t0)
            index.append(' ')
            temp.append(' ')
    if t.data.userNotes is not None:
            # User notes
            index.append('User notes')
            temp.append('t.data.userNotes')
            index.append(' ')
            temp.append(' ')
    statistics = t.getGeneralStatistics()
    for description, (value, precision) in statistics.items():
        if value is not None:
            index.append(description)
            temp.append(float(f'{value:{precision}}'))
    for key, value in t.data.optimizationMessages.items():
        if key == 'Relative projected gradient':
            index.append(key)
            temp.append(float(f'{value:.7g}'))

        else:
            index.append(key)
            try:
                temp.append(float(f'{value}'))
            except:
                temp.append(f'{value}')
    index.append(' ')
    temp.append(' ')
    index.append('Smallest eigenvalue')
    temp.append(float(f'{t.data.smallestEigenValue:.6g}'))
    index.append('Largest eigenvalue')
    temp.append(float(f'{t.data.largestEigenValue:.6g}'))
    index.append('Condition number')
    temp.append(float(f'{t.data.conditionNumber:.6g}'))
    name = 'Report from biogeme ' + bv.getVersion() + ' [' + bv.versionDate + ']'
    temp1 = pd.DataFrame(temp, columns=[name], index=index)
    temp1.to_excel(writer, sheet_name='Estimation report')
    
    # Estimated parameters
    t.getEstimatedParameters(onlyRobust).to_excel(writer, 
                                                  sheet_name='Estimated parameters')
    
    # Correlation of coefficients
    temp2 = t.getCorrelationResults()
    if onlyRobust:
        temp2.drop(columns=['Covariance', 'Correlation', 't-test', 'p-value'], 
                   inplace = True)
    temp2.to_excel(writer, sheet_name='Correlation of coefficients')
    try:
        writer._save()
    except:
        writer.save()
    del writer
    reset_col(fileplace)


def reset_col(filename):
    import biogeme.version as bv
    from openpyxl.utils import get_column_letter 
    from openpyxl.styles import Font, Border, Side, Alignment
    from openpyxl import load_workbook
    wb=load_workbook(filename)
    border_set = Border(left=Side(style=None),
                        right=Side(style=None))
    font = Font(name="Times New Roman", size=16, color='000000', bold=True, 
            italic=False, underline='single', strike=False)
    font2 = Font(bold=False)
    font3 = Font(color='0000FF', underline='single')
    for sheet in wb.sheetnames:
        ws=wb[sheet]
        df=pd.read_excel(filename,sheet).fillna('-')
        df.loc[len(df)]=list(df.columns)
        temp = 0
        if sheet == 'Estimation report':
            ws.cell(row=1, column=2).font = font
            ws.row_dimensions[1].height = 50
            temp = 1
            for i in range(3,9):
                ws.cell(row=i, column=2).font = font3
            ws.cell(row=3, column=2).hyperlink = 'https://www.python.org/'
            ws.cell(row=4, column=2).hyperlink = bv.URL_BIOGEME
            ws.cell(row=5, column=2).hyperlink = bv.URL_FORUM
            ws.cell(row=6, column=2).hyperlink = bv.URL_AUTHOR
            ws.cell(row=7, column=2).hyperlink = bv.URL_DEPARTMENT
            ws.cell(row=8, column=2).hyperlink = bv.URL_UNIVERSITY
            
        for col in df.columns:				
            index=list(df.columns).index(col)					
            letter=get_column_letter(index+1)					
            collen=df[col].apply(lambda x:len(str(x).encode())).max()	
            ws.column_dimensions[letter].width=min((collen*1+3), (collen*1.15))
            if temp == 1:
                if index == 0:
                    for i in range(df.shape[0]):
                        if i <= 7:
                            ws[letter + str(i+2)].border = border_set
                            ws[letter + str(i+1)].alignment = Alignment(horizontal='right', vertical='center')
                            ws.cell(row=i+1, column=1).font = font2
                        else:
                            ws[letter + str(i+1)].alignment = Alignment(horizontal='right', vertical='center')
                            if df[col][i] == ' ':
                                ws[letter + str(i+2)].border = border_set
            
                else:
                    ws[letter + str(1)].alignment = Alignment(horizontal='center', vertical='center')
                    for i in range(1,df.shape[0]):
                        if i <= 7:
                            ws[letter + str(i+1)].alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            ws[letter + str(i+1)].alignment = Alignment(horizontal='right', vertical='center')
                        
            else:
                if index == 0:
                    for i in range(df.shape[0]):
                        ws[letter + str(i+1)].alignment = Alignment(horizontal='left', vertical='center')
                        if df[col][i] == ' ':
                            ws[letter + str(i+2)].border = border_set
                else:
                    ws[letter + str(1)].alignment = Alignment(horizontal='center', vertical='center')
                    for i in range(1,df.shape[0]):
                        ws[letter + str(i+1)].alignment = Alignment(horizontal='right', vertical='center')
                        ws[letter + str(i+1)].number_format = '0.00E+00'
    
    wb.save(filename)

    
def getXlsxHeader():
    import biogeme.version as bv
    """Prepare the header for the HTML file, containing comments and the
    version of Biogeme.

    :return: string containing the header.
    :rtype: str
    """
    a = []
    b = []
    
    a.append(' ')
    b.append('Python package')
    
    a.append('Home page:')
    b.append(bv.URL_BIOGEME)
    
    a.append('Submit questions to:')
    b.append(bv.URL_FORUM)
    
    a.append(' ')
    b.append(bv.AUTHOR)
    a.append(' ')
    b.append(bv.DEPARTMENT)
    a.append(' ')
    b.append(bv.UNIVERSITY)
    return a,b

def resetCol(filename):
    from openpyxl.utils import get_column_letter 
    from openpyxl import load_workbook
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws=wb[sheet]
        df=pd.read_excel(filename,sheet).fillna('-')
        df.loc[len(df)]=list(df.columns)						#把标题行附加到最后一行
        for col in df.columns:				
            index=list(df.columns).index(col)					#列序号
            letter=get_column_letter(index+1)					#列字母
            collen=df[col].apply(lambda x:len(str(x).encode())).max()	#获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width=min((collen*1+3), (collen*1.15))	#也就是列宽为最大长度*1.2 可以自己调整
    wb.save(filename)
    
